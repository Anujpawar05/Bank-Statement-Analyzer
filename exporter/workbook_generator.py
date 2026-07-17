from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


class WorkbookGenerator:
    """
    Generate the final Excel workbook.
    """

    def _style_headers(self, sheet):
        """
        Apply professional styling to header row.
        """

        fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78",
        )

        font = Font(
            bold=True,
            color="FFFFFF",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    def _auto_fit(self, sheet):
        """
        Automatically resize columns.
        """

        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value)),
                        )
                except Exception:
                    pass

            sheet.column_dimensions[column_letter].width = min(
                max_length + 3,
                50,
            )

    def _freeze(self, sheet):
        """
        Freeze header row.
        """

        sheet.freeze_panes = "A2"

    def _filter(self, sheet):
        """
        Enable Excel filters.
        """

        sheet.auto_filter.ref = sheet.dimensions

    def generate(self, result: dict) -> Workbook:
        """
        Generate Excel workbook.
        """

        workbook = Workbook()

        # ==========================================================
        # Transactions Sheet
        # ==========================================================

        sheet = workbook.active
        sheet.title = "Transactions"

        headers = [
            "Date",
            "Description",
            "Debit",
            "Credit",
            "Balance",
            "Category",
            "Merchant",
        ]

        for column, header in enumerate(headers, start=1):
            sheet.cell(
                row=1,
                column=column,
            ).value = header

        row = 2

        for tx in result.get("transactions", []):

            sheet.cell(row=row, column=1).value = tx.get("date")
            sheet.cell(row=row, column=2).value = tx.get("description")
            sheet.cell(row=row, column=3).value = tx.get("debit")
            sheet.cell(row=row, column=4).value = tx.get("credit")
            sheet.cell(row=row, column=5).value = tx.get("balance")
            sheet.cell(row=row, column=6).value = tx.get("category")
            sheet.cell(row=row, column=7).value = tx.get("merchant")

            row += 1

        self._style_headers(sheet)
        self._freeze(sheet)
        self._filter(sheet)
        self._auto_fit(sheet)

        # ==========================================================
        # Summary Sheet
        # ==========================================================

        summary = workbook.create_sheet("Summary")

        summary["A1"] = "Metric"
        summary["B1"] = "Value"

        analysis = result.get("analysis", {})
        financial = analysis.get("financial_summary", {})
        metadata = result.get("metadata", {})

        opening_balance = metadata.get("opening_balance")
        closing_balance = metadata.get("closing_balance")

        total_credit = sum(
            tx.get("credit") or 0
            for tx in result.get("transactions", [])
        )

        total_debit = sum(
            tx.get("debit") or 0
            for tx in result.get("transactions", [])
        )

        rows = [
            ("Bank", result.get("bank_name")),
            ("Account Holder", metadata.get("account_holder")),
            ("Account Number", metadata.get("account_number")),
            ("Statement Period", metadata.get("statement_period")),
            ("", ""),
            ("Opening Balance", opening_balance),
            ("Closing Balance", closing_balance),
            ("", ""),
            ("Total Credits", total_credit),
            ("Total Debits", total_debit),
            ("", ""),
            ("Total Income", financial.get("total_income")),
            ("Total Expense", financial.get("total_expense")),
            ("Net Cash Flow", financial.get("net_cash_flow")),
            ("", ""),
            ("Transactions", len(result.get("transactions", []))),
        ]

        row = 2

        for metric, value in rows:
            summary.cell(row=row, column=1).value = metric
            summary.cell(row=row, column=2).value = value
            row += 1

        self._style_headers(summary)
        self._auto_fit(summary)

        return workbook

    def save(
        self,
        result: dict,
        output_path: str | Path,
    ) -> str:
        """
        Generate and save workbook.
        """

        workbook = self.generate(result)

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(output_path)

        return str(output_path)